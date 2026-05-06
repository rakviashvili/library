"""
Run this script after updating database.xlsx to regenerate data.js.
Usage: python generate_data.py
Requires: pip install openpyxl
"""
import json
try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    exit(1)

wb = openpyxl.load_workbook('database.xlsx', data_only=True)

def cell(sheet, row, col):
    v = sheet.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ''

# main page
s = wb['main page']
main_page = {
    'name':     cell(s, 2, 1),
    'heroText': cell(s, 2, 2),
    'mission':  cell(s, 2, 3),
}

# about
s = wb['ჩვენს შესახებ']
about_text = cell(s, 3, 1)

# events — skip rows with empty name / date / place
s = wb['events']
events = []
for row in range(3, s.max_row + 1):
    num        = cell(s, row, 1)
    name       = cell(s, row, 2)
    date       = cell(s, row, 3)
    place      = cell(s, row, 4)
    annotation = cell(s, row, 5)
    if name and date and place:
        events.append({
            'num':        int(num) if num.isdigit() else 0,
            'name':       name,
            'date':       date,
            'place':      place,
            'annotation': annotation,
        })

# partners — skip placeholder rows
s = wb['Partners']
partners = []
PLACEHOLDER = 'ჯერ არავინ'
for row in range(2, s.max_row + 1):
    num  = cell(s, row, 1)
    name = cell(s, row, 2)
    site = cell(s, row, 3)
    if name and name not in (PLACEHOLDER, 'Partners name'):
        partners.append({
            'num':  int(num) if num.isdigit() else 0,
            'name': name,
            'site': site,
        })

js = """\
// Auto-generated from database.xlsx — do not edit manually.
// Run:  python generate_data.py

const mainPage = {json_main};

const aboutData = {{ text: {json_about} }};

const events = {json_events};

const partners = {json_partners};
""".format(
    json_main    = json.dumps(main_page,  ensure_ascii=False, indent=2),
    json_about   = json.dumps(about_text, ensure_ascii=False),
    json_events  = json.dumps(events,     ensure_ascii=False, indent=2),
    json_partners= json.dumps(partners,   ensure_ascii=False, indent=2),
)

with open('data.js', 'w', encoding='utf-8') as f:
    f.write(js)

print('data.js generated successfully.')
